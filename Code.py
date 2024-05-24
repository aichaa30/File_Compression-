
from openpyxl import load_workbook

bin_to_char = {}
char_to_bin = {}
textFile = "list.txt"


# this function is our function to load the excel sheet into a dictionary.
# Takes true or false, either for binary to decimal or decimal to binary
def loadSheet(text_to_bin):
    wb = load_workbook('Table.xlsx')
    if text_to_bin:  # if going from txt to binary, get letter as key and binary as value
        for row in wb.worksheets[0].iter_rows(min_row=2, max_row=81, max_col=2):
            char_to_bin[str(row[0].value)] = row[1].value
        return char_to_bin
    else:  # if binary to decimal, get binary as key and letter as value
        for row in wb.worksheets[0].iter_rows(min_row=2, max_row=81, max_col=2):
            bin_to_char[str(row[1].value)] = row[0].value
        return bin_to_char


# this is our main code function
# takes filename as input, outputs a file with decimal number of bits and binary value of text.
def Code(filename):
    table = loadSheet(True)
    file = open(filename, 'r')
    output = open("BinOutput.txt", 'w+')
    character = file.read()
    bin = ""
    i = 0
    while i < len(character):  # iterates through the string of text
        if character[i] == '\n':  # checks if there is a new line
            bin += table["newline"]
            i += 1
        elif character.find("week") == i:  # these next if statements check for words we have in our table
            bin += table["week"]
            i += 4
        elif character.find("the") == i:
            bin += table["the"]
            i += 3
        elif character.find("and") == i:
            bin += table["and"]
            i += 3
        elif character.find("it") == i:
            bin += table["it"]
            i += 2
        elif character.find("is") == i:
            bin += table["is"]
            i += 2
        elif character.find("to") == i:
            bin += table["to"]
            i += 2
        elif character.find("of") == i:
            bin += table["of"]
            i += 2
        else:  # if not a word, just get the character
            bin += table[character[i]]
            i += 1
    output.write((str(len(bin)) + "." + bin))
    file.close()
    output.close()


Code(textFile)
